import html
import io
import re
from pathlib import Path

import streamlit as st
from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from config_store import ConfigStore
from jira_client import JiraClient
from llm_client import LLMClient

load_dotenv()
ROOT = Path(__file__).parent
TEMPLATES = ROOT / 'templates'
OUTPUTS = ROOT / 'outputs'
OUTPUTS.mkdir(exist_ok=True)

st.set_page_config(page_title='Jira Test Case Generator', layout='wide')

config = ConfigStore(config_path=ROOT / 'config.json')

JIRA_KEY_RE = re.compile(r'([A-Z][A-Z0-9]+-\d+)')
FENCE_RE = re.compile(r'```[a-zA-Z]*\s*\n(.*?)```', re.DOTALL)
EXPECTED_COLUMNS = ['Test ID', 'Description', 'Pre-conditions', 'Steps', 'Expected Result', 'Priority']
SECRET_KEYS = ('jira_api_token', 'groq_api_key')

DEFAULT_TEMPLATE = (
    'ROLE: You are a Senior QA Engineer.\n\n'
    'TASK: Write test cases for {{jira_id}} using only the requirement below.\n\n'
    'Return exactly one Markdown table with these columns and nothing else:\n'
    '| Test ID | Description | Pre-conditions | Steps | Expected Result | Priority |\n\n'
    'TITLE: {{summary}}\n\nACCEPTANCE CRITERIA:\n{{acceptance_criteria}}\n\n'
    'DESCRIPTION:\n{{description}}\n'
)


def extract_jira_key(text):
    match = JIRA_KEY_RE.search(text or '')
    return match.group(1) if match else None


def load_template():
    path = TEMPLATES / 'test_case_template.md'
    return path.read_text(encoding='utf-8') if path.exists() else DEFAULT_TEMPLATE


def build_prompt(jira_key, issue):
    return (
        load_template()
        .replace('{{jira_id}}', jira_key or 'the requirement')
        .replace('{{summary}}', issue.get('summary', '') or 'Not specified')
        .replace('{{acceptance_criteria}}', issue.get('acceptance_criteria', '') or 'Not specified')
        .replace('{{description}}', issue.get('description', '') or 'Not specified')
        .replace('{{preconditions}}', '- The user has access to the application under test')
    )


def _is_separator_row(line):
    stripped = line.strip()
    return '-' in stripped and re.fullmatch(r'[|\-:\s]+', stripped) is not None


def extract_table(text):
    """Collect the pipe-delimited rows from the response, ignoring blank lines between them."""
    rows = []
    for line in (text or '').splitlines():
        stripped = line.strip()
        if stripped.startswith('|') and stripped.count('|') >= 3:
            rows.append(stripped)
    return rows if len(rows) >= 2 else []


def _cells(row):
    return [c.strip() for c in row.strip().strip('|').split('|')]


def _row(cells):
    return '| ' + ' | '.join(cells) + ' |'


def _fit(cells, width):
    """Coerce a row to the required width, assuming trailing columns are the reliable ones."""
    if len(cells) == width:
        return cells
    if len(cells) > width:
        return cells[:width - 1] + [' '.join(cells[width - 1:])]
    if len(cells) >= width - 1:
        head, middle, tail = cells[:1], cells[1:-3], cells[-3:]
        gap = width - 4 - len(middle)
        return head + middle + ['Not specified'] * gap + tail
    return cells + ['Not specified'] * (width - len(cells))


def table_records(rows):
    """Force the response into the required six-column shape and drop repeated cases."""
    if not rows:
        return []
    width = len(EXPECTED_COLUMNS)
    first = _cells(rows[0])
    has_header = any('test id' in c.lower() or 'description' in c.lower() for c in first)
    body = []
    seen = set()
    for row in rows[1:] if has_header else rows:
        if _is_separator_row(row):
            continue
        cells = _fit(_cells(row), width)
        if not any(cells):
            continue
        key = cells[1].lower().rstrip('.')
        if key and key in seen:
            continue
        seen.add(key)
        cells[0] = f'TC-{len(body) + 1:03d}'
        body.append(cells)
    return body


def records_to_markdown(records):
    if not records:
        return ''
    width = len(EXPECTED_COLUMNS)
    separator = '|' + '|'.join([' --- '] * width) + '|'
    return '\n'.join([_row(EXPECTED_COLUMNS), separator] + [_row(cells) for cells in records])


def _excel_cell(value):
    text = (value or '').replace('<br>', '\n').replace('<br/>', '\n').replace('<br />', '\n')
    return re.sub(r'\s+\n', '\n', text).strip()


def build_workbook(records, jira_key=None, issue=None):
    wb = Workbook()
    sheet = wb.active
    sheet.title = 'Test Cases'

    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill('solid', fgColor='1F4E79')
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell_align = Alignment(vertical='top', wrap_text=True)
    thin = Border(
        left=Side(style='thin', color='D0D7DE'),
        right=Side(style='thin', color='D0D7DE'),
        top=Side(style='thin', color='D0D7DE'),
        bottom=Side(style='thin', color='D0D7DE'),
    )

    for col, heading in enumerate(EXPECTED_COLUMNS, start=1):
        cell = sheet.cell(1, col, heading)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin

    for row_index, values in enumerate(records, start=2):
        for col, value in enumerate(values, start=1):
            cell = sheet.cell(row_index, col, _excel_cell(value))
            cell.alignment = cell_align
            cell.border = thin
        sheet.row_dimensions[row_index].height = 60

    widths = [12, 42, 32, 55, 42, 12]
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width
    sheet.auto_filter.ref = sheet.dimensions
    sheet.freeze_panes = 'A2'
    sheet.row_dimensions[1].height = 22

    meta = wb.create_sheet('Ticket')
    meta.append(['Field', 'Value'])
    meta['A1'].font = header_font
    meta['B1'].font = header_font
    meta['A1'].fill = header_fill
    meta['B1'].fill = header_fill
    issue = issue or {}
    for label, value in (
        ('Jira key', jira_key or issue.get('key') or 'ad-hoc'),
        ('Summary', issue.get('summary') or ''),
        ('Acceptance criteria', issue.get('acceptance_criteria') or ''),
    ):
        meta.append([label, value])
    meta.column_dimensions['A'].width = 24
    meta.column_dimensions['B'].width = 80
    for row in meta.iter_rows(min_row=1, max_row=meta.max_row, min_col=1, max_col=2):
        for cell in row:
            cell.alignment = Alignment(vertical='top', wrap_text=True)

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def to_display(markdown_text):
    """Escape the model output, then re-allow only <br> so numbered steps break onto their own lines."""
    escaped = html.escape(markdown_text or '', quote=False)
    return re.sub(r'&lt;\s*br\s*/?\s*&gt;', '<br>', escaped, flags=re.IGNORECASE)


def format_response(text):
    """Return (markdown, is_table, records). Falls back to the raw text when no table was produced."""
    cleaned = (text or '').strip()
    fenced = FENCE_RE.search(cleaned)
    if fenced and '|' in fenced.group(1):
        cleaned = fenced.group(1).strip()
    records = table_records(extract_table(cleaned))
    if records:
        return records_to_markdown(records), True, records
    return cleaned, False, []


def generate(chat_input, settings):
    jira_key = extract_jira_key(chat_input)
    issue = None
    if jira_key:
        if not settings.get('jira_url'):
            st.error('Jira settings missing. Open the settings page in the sidebar.')
            return None
        try:
            issue = JiraClient(settings).get_issue(jira_key)
        except Exception as e:
            st.error(f'Jira fetch failed: {e}')
            return None

    prompt = build_prompt(jira_key, issue) if issue else chat_input
    with st.spinner(f'Generating test cases with {settings.get("llm_provider", "ollama")}...'):
        response = LLMClient(settings).send_prompt(prompt)

    markdown, is_table, records = format_response(response)
    return {
        'jira_key': jira_key,
        'prompt': prompt,
        'raw': response,
        'markdown': markdown,
        'is_table': is_table,
        'records': records,
        'issue': issue,
    }


st.title('Jira Test Case Generator')
settings = config.get_settings() or {}

col1, col2 = st.columns([3, 1])

with col1:
    st.header('Chat')
    chat_input = st.text_area('Message', height=120, placeholder='create test cases for KAN-2')
    if st.button('Send', type='primary'):
        if not chat_input.strip():
            st.warning('Please enter a message or Jira key')
        else:
            if not extract_jira_key(chat_input):
                st.info('No Jira key detected. Sending the message straight to the LLM.')
            result = generate(chat_input, settings)
            if result:
                st.session_state['result'] = result

    result = st.session_state.get('result')
    if result:
        st.subheader('Generated Test Cases')
        if not result['is_table']:
            st.warning(
                'The model did not return a table. Showing the raw response — '
                'retry, or switch to a larger model in settings.'
            )
        st.markdown(to_display(result['markdown']), unsafe_allow_html=True)

        stem = f"{result['jira_key'] or 'ad-hoc'}_test_cases"
        md_name = f'{stem}.md'
        xlsx_name = f'{stem}.xlsx'
        xlsx_bytes = None
        if result.get('records'):
            xlsx_bytes = build_workbook(result['records'], result.get('jira_key'), result.get('issue'))

        download_col, save_col = st.columns(2)
        with download_col:
            st.download_button(
                'Download .md',
                data=result['markdown'],
                file_name=md_name,
                mime='text/markdown',
                key='download_md',
            )
            if xlsx_bytes:
                st.download_button(
                    'Download .xlsx',
                    data=xlsx_bytes,
                    file_name=xlsx_name,
                    mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                    key='download_xlsx',
                )
            else:
                st.caption('Excel export needs a table. Retry generation if only free text was returned.')
        with save_col:
            if st.button('Save to outputs'):
                saved = []
                md_path = OUTPUTS / md_name
                md_path.write_text(result['markdown'], encoding='utf-8')
                saved.append(str(md_path))
                if xlsx_bytes:
                    xlsx_path = OUTPUTS / xlsx_name
                    xlsx_path.write_bytes(xlsx_bytes)
                    saved.append(str(xlsx_path))
                st.success('Saved:\n' + '\n'.join(saved))

        with st.expander('Ticket content sent to the model'):
            st.code(result['prompt'])
        with st.expander('Raw model response'):
            st.code(result['raw'])

with col2:
    st.header('Quick Actions')
    st.caption("Use the 'settings' page in the sidebar to configure credentials.")
    st.markdown('---')
    st.write('Saved settings:')
    st.json({
        key: ('********' if key in SECRET_KEYS and value else value)
        for key, value in (settings or {}).items()
    })
